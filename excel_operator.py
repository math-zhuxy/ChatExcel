from zhipuai import ZhipuAI
import json
from openpyxl import load_workbook
from openpyxl.cell import Cell
import time
from typing import List
from typing import Tuple
from utils import OUTPUT_STATE

class GLM_MODEL:
    def __init__(self):      
        self.FILE_PATH = ""
        # self.SAVE_FILE_PATH = "./test.xlsx"
        try:
            with open("config.json", 'r', encoding='utf-8') as file:
                config_data = json.load(file)
                api_key = config_data["api_key"]
        except FileNotFoundError:
            print("Error: The file config.json was not found.")
        except json.JSONDecodeError:
            print("Error: The file config.json is not a valid JSON file.")
        except KeyError:
            print("Error: The 'api_key' was not found in the JSON data.")


        self.CHATGLM_API_KEY = api_key

        self.chatglm_client = ZhipuAI(api_key=self.CHATGLM_API_KEY)

        self.FUNCTION_METHOD = [
            {
                "type": "function",
                "function": {
                    "name": "write_excel",
                    "description": "将数据存放到Excel文件中指定的工作表中指定的单元格内",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "sheet":{
                                "type": "string",
                                "description": "要存储的工作表名称，例如 'sheet1'。"
                            },
                            "cell_range": {
                                "type": "string",
                                "description": "要存储的单元格，例如 'C11'。"
                            },
                            "data":{
                                "type": "string",
                                "description": "要存储的数据，例如 'hello world'。"
                            }
                        },
                        "required": ["sheet","cell_range","data"],
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "read_excel",
                    "description": "读取Excel文件中指定的工作表中指定的单元格范围的数据",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "sheet":{
                                "type": "string",
                                "description": "要读取的工作表名称，例如 'Sheet1'。"
                            },
                            "cell_range": {
                                "type": "string",
                                "description": "要读取的单元格范围，例如 'A1:B10'。"
                            }
                        },
                        "required": ["sheet","cell_range"],
                    }
                }
            },
        ]
        self.messages=[
            {
            "role": "system",
            "content":  """
                            1.你是一个Excel的工作助手，可以帮助用户完成Excel的相关操作
                            2.你的输出语言必须是中文。
                            3.涉及对Excel的操作时尽量调用read_excel或write_excel函数
                            4.不要假设或猜测传入函数的参数值。如果用户的描述不明确，请要求用户提供必要信息
                        """
            }
        ]

    def read_excel(self, sheet: str, cell_range: str) -> List[List[str]]:
        print("function read excel called")
        try:
            ws = load_workbook(filename=self.FILE_PATH, data_only=True)[sheet]  # 确保我们得到的是值而不是公式
        except FileExistsError:
            print("文件名或工作簿不存在")
            return [[]]
        if ':' not in cell_range:
            cell_range=f"{cell_range}:{cell_range}"
        # 读取指定范围的数据
        data = ws[cell_range]

        if isinstance(data,Cell):
            data = [[data]]

        # 将数据转换为列表的列表，并尝试将所有元素转换为整数；如果为空则设为0
        result = []
        for row in data:
            int_row = []
            for cell in row:
                cell_value = cell.value
                try:
                    # 检查单元格是否有内容，如果没有则设为''
                    if cell_value is None or str(cell_value).strip() == '':
                        str_cell = ''
                    else:
                        str_cell = str(cell_value)
                except (ValueError, TypeError) as e:
                    print(f"Error converting cell value '{cell_value}' to integer: {e}")
                    str_cell = ''  # 转换失败也设为空字符串
                int_row.append(str_cell)
            result.append(int_row)
        print("function read excel called succssfully")
        return result
    
    def is_number(self, num: str) -> bool:
        try:
            float(num)
            return True
        except ValueError:
            return False
    
    def write_excel(self, sheet: str, cell_range: str, data: str) -> str:
        print("function write excel called")
        try:
            wb = load_workbook(filename=self.FILE_PATH, data_only=True)
            if self.is_number(data):
                wb[sheet][cell_range] = float(data)
            else:
                wb[sheet][cell_range] = data
            wb.save(filename=self.FILE_PATH)
        except Exception:
            print(f"发生错误{Exception}")
            return "fail"

        print("function write excel called succssfully")
        return "存储数据成功"

    def test_readexcel(self, test_sheet: str, test_range: str) -> None:
        print(self.read_excel(test_sheet, test_range))

glm_model = GLM_MODEL()

def excel_operate(user_input: str, file_path: str,function_called: str ,progress_callback) -> Tuple[str, OUTPUT_STATE]:
    print("start excel_operator function")
    glm_model.FILE_PATH = file_path

    glm_model.messages.append({
        "role": "user",
        "content": user_input
    })

    tool_choice = "auto"

    if function_called == "none":
        tool_choice = ""
    elif function_called == "auto":
        tool_choice = "auto"
    elif function_called == "read":
        tool_choice = {"type": "function", "function": {"name": "read_excel"}}
    elif function_called == "write":
        tool_choice = {"type": "function", "function": {"name": "write_excel"}}
    else:
        return "参数function_called不正确", OUTPUT_STATE.INCORRECT_PARAMETER
    
    # 创建完请求
    for i in range(11):
        progress_callback(i, "正在初步分析需求")
        time.sleep(0.01)

    response = glm_model.chatglm_client.chat.completions.create(
        model = "glm-4-flash", 
        messages = glm_model.messages,
        tools = glm_model.FUNCTION_METHOD,
        tool_choice = tool_choice
    )

    glm_model.messages.append(response.choices[0].message.model_dump())

    # print(response)

    # 获取到函数调用信息
    for i in range(35):
        progress_callback(i+11, "正在获取模型信息")
        time.sleep(0.01)

    while response.choices[0].message.tool_calls:
        tool_call = response.choices[0].message.tool_calls[0]
        args = tool_call.function.arguments
        function_result = {}

        if tool_call.function.name == "read_excel":
            function_result = glm_model.read_excel(**json.loads(args))
        if tool_call.function.name == "write_excel":
            function_result = glm_model.write_excel(**json.loads(args))
    
        if tool_call.function.name == "read_excel" and function_result == [[]]:
            glm_model.messages = []
            for i in range(55):
                progress_callback(i+46, "函数调用错误")
                time.sleep(0.005)
            return "文件路径或工作表名称出错，已清空消息列表", OUTPUT_STATE.FILE_PATH_ERROR

        if tool_call.function.name == "write_excel" and function_result == "fail":
            glm_model.messages = []
            for i in range(55):
                progress_callback(i+46, "函数调用错误")
                time.sleep(0.005)
            return "文件路径或工作表名称出错，已清空消息列表", OUTPUT_STATE.FILE_PATH_ERROR
        
        # 调用完函数
        for i in range(20):
            progress_callback(i+46, "正在调用函数")
            time.sleep(0.05)
        
        glm_model.messages.append({
            "role": "tool",
            "content": f"{json.dumps(function_result)}",
            "tool_call_id":tool_call.id
        })

        response = glm_model.chatglm_client.chat.completions.create(
            model = "glm-4-flash",  
            messages = glm_model.messages,
            tools = glm_model.FUNCTION_METHOD,
            tool_choice = tool_choice
        )

        glm_model.messages.append(response.choices[0].message.model_dump())

        # print(response)

        # 获取到最终请求
        for i in range(35):
            progress_callback(i+66, "模型正在解析结果")
            time.sleep(0.005)
        return response.choices[0].message.content, OUTPUT_STATE.FUNCTION_CALLED_SUCCESS
    else:
        # 无需调用函数
        for i in range(90):
            progress_callback(i+11, "无需调用函数，模型正在整理信息")
    return response.choices[0].message.content, OUTPUT_STATE.MESSAGE_SUCCESS
    